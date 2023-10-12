from openpyxl import load_workbook
from yattag import Doc, indent

from tkinter import Tk

import tkinter.filedialog as fd

import pandas as pd

Tk().withdraw()
print('Выберите файл')
filename = fd.askopenfilename()
if filename == '':
    print('Не выбран файл EXCEL')

print('Выберите путь для сохранения штрихкодов')
directory = fd.askdirectory()
if directory == '':
    print('Не выбран путь для сохранения штрихкодов')

row_min = int(input('Enter first row: '))
row_max = int(input('Enter end row: '))
col_min = int(input('Enter first col: '))
col_max = int(input('Enter end col: '))

df = pd.read_excel(filename)

# Load our Excel File
wb = load_workbook(filename)
# Getting an object of active sheet 1
ws = wb.worksheets[0]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
doc.asis(xml_header)

with tag('rss', 'xmlns:g="http://base.google.com/ns/1.0" version="2.0"'):
    with tag('channel'):
        with tag('title'):
            text('fb_ae')
        with tag('link'):
            text('https://url.ru')
        with tag('description'):
            text('')
        for row in ws.iter_rows(min_row=row_min, max_row=row_max, min_col=col_min, max_col=col_max):
            row = [cell.value for cell in row]
            with tag('item'):
                if row[0] != None:
                    with tag("g:id"):
                        text(row[0])
                if row[1] != None:
                    with tag("g:title"):
                        text(row[1])
                if row[2] != None:
                    with tag("g:description"):
                        txt = ''
                        for i in row[2]:
                            if i == '\n':
                                i = ' '
                            txt += i
                        text(txt)
                if row[3] != None:
                    with tag("g:link"):
                        text(row[3])
                if row[4] != None:
                    with tag("g:image_link"):
                        text(row[4])
                if row[5] != None:
                    with tag("g:additional_image_link"):
                        text(row[5])
                if row[6] != None:
                    with tag("g:availability"):
                        text(row[6])
                if row[7] != None:
                    with tag("g:price"):
                        text(row[7])
                if row[8] != None:
                    with tag("g:sale_price"):
                        text(row[8])
                if row[9] != None:
                    with tag("g:google_product_category"):
                        text(row[9])
                if row[10] != None:
                    with tag("g:product_type"):
                        text(row[10])
                if row[11] != None:
                    with tag("g:brand"):
                        text(row[11])
                if row[12] != None:
                    with tag("g:gtin"):
                        text(row[12])
                if row[13] != None:
                    with tag("g:condition"):
                        text(row[13])
                if row[14] != None:
                    with tag("g:material"):
                        text(row[14])
                if row[15] != None:
                    with tag("g:color"):
                        text(row[15])



result = indent(
    doc.getvalue(),
    indentation='   ',
    # newline = '\r\n',
    indent_text=False
)

with open(directory + '/fb_ae.xml', "w", encoding="utf-8") as f:
    f.write(result)
